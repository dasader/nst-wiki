"""Stage 0 포맷 분기 + MD/XLSX 파싱. PDF는 parse_pdf 모듈로 위임."""
import json
from pathlib import Path


def run_pipeline(source_dir: Path) -> None:
    original = next(source_dir.glob("original.*"), None)
    if original is None:
        raise ValueError(f"no original.* file in {source_dir}")
    out = source_dir / "parsed"
    out.mkdir(exist_ok=True)
    ext = original.suffix.lower()
    if ext == ".pdf":
        from pipeline.parse_pdf import parse_pdf  # 무거운 파싱 의존은 지연 임포트

        parse_pdf(original, out)
    elif ext == ".md":
        parse_md(original, out)
    elif ext == ".xlsx":
        parse_xlsx(original, out)
    else:
        raise ValueError(f"unsupported format: {ext}")


def _write_chunks(out: Path, chunks: list[dict]) -> None:
    (out / "chunks.json").write_text(
        json.dumps(chunks, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def chunk_markdown(md: str) -> list[str]:
    """마크다운을 헤딩(#) 경계로 분할한 비어있지 않은 섹션 리스트."""
    sections, cur = [], []
    for line in md.splitlines():
        if line.startswith("#") and cur:
            sections.append("\n".join(cur).strip())
            cur = []
        cur.append(line)
    if cur:
        sections.append("\n".join(cur).strip())
    return [s for s in sections if s]


def parse_md(src: Path, out: Path) -> None:
    text = src.read_text(encoding="utf-8")
    (out / "document.md").write_text(text, encoding="utf-8")
    chunks = [
        {"id": f"c{i:03d}", "type": "text", "page": None, "text": s}
        for i, s in enumerate(chunk_markdown(text), 1)
    ]
    _write_chunks(out, chunks)


def _header(cells: tuple) -> list[str]:
    """첫 행 → 컬럼명. 빈 칸은 Unnamed: N, 중복은 .1/.2 접미 (pandas read_excel과 동일 규약).

    중복 컬럼명을 그대로 두면 LLM 표 매핑이 두 열을 구분하지 못한다.
    """
    cols: list[str] = []
    for j, c in enumerate(cells):
        name = str(c) if c is not None else f"Unnamed: {j}"
        dup = sum(1 for prev in cols if prev == name or prev.startswith(f"{name}."))
        cols.append(f"{name}.{dup}" if dup else name)
    return cols


def parse_xlsx(src: Path, out: Path) -> None:
    from openpyxl import load_workbook

    tables_dir = out / "tables"
    tables_dir.mkdir(exist_ok=True)
    chunks = []
    wb = load_workbook(src, read_only=True, data_only=True)  # data_only: 수식 대신 캐시된 값
    for i, ws in enumerate(wb.worksheets, 1):
        rows = ws.iter_rows(values_only=True)
        cols = _header(next(rows, ()))
        body = [
            list(r[:len(cols)]) + [None] * max(0, len(cols) - len(r))  # 헤더 폭에 맞춤
            for r in rows
            if any(c is not None for c in r)  # 완전 공백 행은 버린다
        ]
        ref = f"tables/table_{i:03d}.json"
        payload = {"table_title": ws.title, "columns": cols, "rows": body}
        (out / ref).write_text(
            json.dumps(payload, ensure_ascii=False, default=str), encoding="utf-8"
        )
        chunks.append({"id": f"c{i:03d}", "type": "table", "page": None, "ref": ref})
    wb.close()
    _write_chunks(out, chunks)
